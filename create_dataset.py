# -*- coding: utf-8 -*-
"""
Created on Sun Nov 28 21:53:45 2021

@author: FaresBougourzi
"""

# Face Beauty Prediction Project
# Bougourzi Fares

import torch
import numpy as np
import math
import cv2
import csv
import os
from skimage import  transform 
import scipy.io as sio

from utils import face_align_dt_land
import tqdm as tqdm 
import matplotlib.pyplot as plt
import openpyxl
import argparse
import struct


def _read_points(land_read):
    with open(land_read, 'rb') as f:
        data = f.read()
        points = struct.unpack('i172f', data)
        return np.reshape(np.array(points)[1:], (-1, 2))


def create_train_test_files():
    parser = argparse.ArgumentParser()
    parser.add_argument('--data_path', type=str, default='./Dataset',
                        help='Path of the dataset')
    parser.add_argument('--Fold', type=int, default=6,
                        help='1-5 stand for the 5 folds and 6 for 60-40 data splitting')    
    opt = parser.parse_args()


    path_scores = os.path.join(opt.data_path, 'SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/All_Ratings.xlsx')
    print(opt.data_path)
    print(path_scores)
    input_data_scores = openpyxl.load_workbook(path_scores)
    data_scores = input_data_scores.worksheets[0]
    
    database_path =  os.path.join(opt.data_path, 'SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/Images')
    land_path =  os.path.join(opt.data_path, 'SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/facial landmark')

    if opt.Fold == 6:        
        # train 60 test 40
        train_labels_path = os.path.join(opt.data_path, 'SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/train_test_files/split_of_60%training and 40%testing/train.txt')
        test_labels_path = os.path.join(opt.data_path, "SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/train_test_files/split_of_60%training and 40%testing/test.txt")
            
        
    else:
        # fold1
        train_labels_path = os.path.join(opt.data_path, 'SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/train_test_files/5_folders_cross_validations_files/cross_validation_1/train_' + str(opt.Fold) +'.txt')
        test_labels_path = os.path.join(opt.data_path, 'SCUT-FBP5500_v2.1/SCUT-FBP5500_v2/train_test_files/5_folders_cross_validations_files/cross_validation_1/test_' + str(opt.Fold) +'.txt')

        
            

    # create training data
    f = open(train_labels_path)
    my_list = f.readlines()
    f.close()
    
    
    Training_data = []
    Training_label = []
    Training_sigma = []
    for line in my_list:
        img_name, label = str.split(line[:-1], ' ')
    
        full_path_image = os.path.join(database_path, img_name)
        img = cv2.imread(full_path_image)
        try:
            land_read = os.path.join(land_path, img_name[:-3] + 'pts')
            vec = _read_points(land_read)
            try: 
                cropped_face = face_align_dt_land(img, vec, (224, 224))
                image = cropped_face.transpose((2, 0, 1))
                Training_data.append(np.array(image))
                Training_label.append(float(label))
                scores = np.zeros([60,1])
                lll = -1
                for row in data_scores.iter_rows():
                    if img_name == row[1].value:
                        lll += 1
                        scores[lll, 0] = row[2].value
    
                Training_sigma.append(float(np.std(scores))) 
    
                
            except:
                print('fail')
        except:
            print('both failed')
            
    X = torch.Tensor([i for i in Training_data]) 
    y = torch.Tensor([i for i in Training_label])
    z = torch.Tensor([i for i in Training_sigma])
    training= (X, y, z)
    
    torch.save(training, 'train_fold'+ str(opt.Fold)+ '.pt')
            
                
    
    
    Test_data = []
    Test_label = []
    Testing_sigma = []
    
    f = open(test_labels_path)
    my_list = f.readlines()
    f.close()
    
    #Testing_data = []
    #Testing_label = []
    for line in my_list:
        img_name, label = str.split(line[:-1], ' ')
        full_path_image = os.path.join(database_path, img_name)
        img = cv2.imread(full_path_image)
        try:
            land_read = os.path.join(land_path, img_name[:-3]+'pts')
            vec = _read_points(land_read)
            try:
                cropped_face = face_align_dt_land(img, vec, (224, 224))
                image = cropped_face.transpose((2, 0, 1))
                Test_data.append(np.array(image))
                Test_label.append(float(label))
                scores = np.zeros([60,1])
                lll = -1
                for row in data_scores.iter_rows():
                    if img_name == row[1].value:
                        lll += 1
                        scores[lll, 0] = row[2].value

                Testing_sigma.append(float(np.std(scores)))

                
            except:
                print(line)
        except:
            print('both failed')
            
                
    X = torch.Tensor([i for i in Test_data]) 
    y = torch.Tensor([i for i in Test_label]) 
    z = torch.Tensor([i for i in Training_sigma])
    training= (X, y, z)
    torch.save(training, 'test_fold'+ str(opt.Fold)+ '.pt')

if __name__ == "__main__":
    create_train_test_files()
